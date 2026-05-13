"""
importar_excel.py
-----------------
Importa ControlInventario 2026.xlsx → PostgreSQL (tablas inventario_*)
Uso:
    python importar_excel.py
    python importar_excel.py --excel "C:/ruta/al/archivo.xlsx"
    python importar_excel.py --host localhost --port 5432 --db outiltech_db --user toyota_user --pass Toyota2026!
"""

import argparse
import socket
import sys
from datetime import date, datetime
from pathlib import Path

# Forzar IPv4 para resolver problemas de DNS IPv6 (Supabase en Windows)
_orig_getaddrinfo = socket.getaddrinfo
def _ipv4_only(host, port, family=0, *args, **kwargs):
    try:
        return _orig_getaddrinfo(host, port, socket.AF_INET, *args, **kwargs)
    except socket.gaierror:
        return _orig_getaddrinfo(host, port, family, *args, **kwargs)
socket.getaddrinfo = _ipv4_only

try:
    import openpyxl
    import psycopg2
    from psycopg2.extras import execute_batch
except ImportError:
    print("Instalando dependencias...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "psycopg2-binary", "-q"])
    import openpyxl
    import psycopg2
    from psycopg2.extras import execute_batch

# ── Configuración por defecto ────────────────────────────────────────────────
DEFAULT_EXCEL = r"C:\Users\user\Downloads\ControlInventario 2026.xlsx"
DEFAULT_DB = {
    "host":     "localhost",
    "port":     5432,
    "dbname":   "outiltech_db",
    "user":     "toyota_user",
    "password": "Toyota2026!",
}

# ── Utilidades ───────────────────────────────────────────────────────────────
def safe_int(v, default=0):
    try: return int(v) if v is not None else default
    except: return default

def safe_decimal(v, default=0.0):
    try: return float(v) if v is not None else default
    except: return default

def safe_str(v, default=""):
    if v is None: return default
    return str(v).strip()

def safe_date(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    return date.today()

# ── Leer Excel ───────────────────────────────────────────────────────────────
def leer_excel(path: str):
    print(f"\n📂 Leyendo: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    stock, entradas, salidas = [], [], []

    # ── STOCK (fila encabezado=4, datos desde fila 5) ────────────────────────
    ws = wb["Stock"]
    for row in ws.iter_rows(min_row=5, values_only=True):
        codigo = safe_str(row[0])
        if not codigo: continue
        stock.append({
            "codigo_producto":              codigo.upper(),
            "descripcion":                  safe_str(row[1]) or codigo,
            "lote":                         safe_str(row[2]) or codigo,
            "entradas":                     safe_int(row[3]),
            "salidas":                      safe_int(row[4]),
            "stock_actual":                 safe_int(row[5]),
            "costo_unitario":               safe_decimal(row[6]),
            "precio_venta":                 safe_decimal(row[7]),
            "importe_inventario_proveedor": safe_decimal(row[8]),
            "venta":                        safe_decimal(row[10]),
        })
    print(f"  ✅ Stock: {len(stock)} productos")

    # ── ENTRADAS (fila encabezado=3, datos desde fila 4) ────────────────────
    ws2 = wb["Entradas"]
    for row in ws2.iter_rows(min_row=4, values_only=True):
        codigo = safe_str(row[2])
        if not codigo: continue
        entradas.append({
            "nro_documento":  safe_int(row[0], 1),
            "fecha":          safe_date(row[1]),
            "codigo_producto": codigo.upper(),
            "descripcion":    safe_str(row[3]),
            "lote":           safe_str(row[4]),
            "cantidad":       safe_int(row[5], 1),
        })
    print(f"  ✅ Entradas: {len(entradas)} registros")

    # ── SALIDAS (fila encabezado=3, datos desde fila 4) ─────────────────────
    ws3 = wb["Salidas"]
    for row in ws3.iter_rows(min_row=4, values_only=True):
        codigo = safe_str(row[2])
        if not codigo or codigo == "-": continue
        salidas.append({
            "nro_documento":  safe_int(row[0], 1),
            "fecha":          safe_date(row[1]),
            "codigo_producto": codigo.upper(),
            "descripcion":    safe_str(row[3]),
            "lote":           safe_str(row[4]),
            "cantidad":       safe_int(row[5], 1),
            "precio_venta":   safe_decimal(row[6]),
            "utilidad":       safe_decimal(row[7]),
        })
    print(f"  ✅ Salidas: {len(salidas)} registros")

    wb.close()
    return stock, entradas, salidas

# ── Importar a PostgreSQL ────────────────────────────────────────────────────
def importar(stock, entradas, salidas, db_cfg):
    host = db_cfg.get("host", "localhost")
    port = db_cfg.get("port", 5432)
    dbname = db_cfg.get("dbname", "postgres")
    print(f"\n🐘 Conectando a PostgreSQL ({host}:{port}/{dbname})...")
    hostaddr = db_cfg.pop("hostaddr", None)
    if hostaddr:
        parts = [f"host={host}", f"hostaddr={hostaddr}", f"port={port}", f"dbname={dbname}",
                 f"user={db_cfg['user']}", f"password={db_cfg['password']}"]
        if db_cfg.get("sslmode"):
            parts.append(f"sslmode={db_cfg['sslmode']}")
        parts.append("connect_timeout=15")
        conn = psycopg2.connect(" ".join(parts))
    else:
        conn = psycopg2.connect(**db_cfg)
    conn.autocommit = False
    cur = conn.cursor()
    cur.execute("SET search_path = public")

    try:
        # ── Insertar Stock (UPSERT por codigo_producto) ──────────────────────
        print("  Importando inventario_stock...")
        sql_stock = """
            INSERT INTO inventario_stock
                (codigo_producto, descripcion, lote, entradas, salidas, stock_actual,
                 costo_unitario, precio_venta, importe_inventario_proveedor, venta)
            VALUES
                (%(codigo_producto)s, %(descripcion)s, %(lote)s, %(entradas)s, %(salidas)s,
                 %(stock_actual)s, %(costo_unitario)s, %(precio_venta)s,
                 %(importe_inventario_proveedor)s, %(venta)s)
            ON CONFLICT (codigo_producto) DO UPDATE SET
                descripcion                  = EXCLUDED.descripcion,
                lote                         = EXCLUDED.lote,
                entradas                     = EXCLUDED.entradas,
                salidas                      = EXCLUDED.salidas,
                stock_actual                 = EXCLUDED.stock_actual,
                costo_unitario               = EXCLUDED.costo_unitario,
                precio_venta                 = EXCLUDED.precio_venta,
                importe_inventario_proveedor = EXCLUDED.importe_inventario_proveedor,
                venta                        = EXCLUDED.venta,
                updated_at                   = NOW()
        """
        execute_batch(cur, sql_stock, stock, page_size=100)
        print(f"    ✅ {len(stock)} productos insertados/actualizados")

        # ── Insertar Entradas (sin triggers para no duplicar stock) ──────────
        # Se inserta con SET session_replication_role = replica para desactivar triggers
        print("  Importando inventario_entradas (sin re-calcular stock)...")
        cur.execute("SET session_replication_role = replica")  # desactiva triggers
        sql_entradas = """
            INSERT INTO inventario_entradas
                (nro_documento, fecha, codigo_producto, descripcion, lote, cantidad)
            VALUES
                (%(nro_documento)s, %(fecha)s, %(codigo_producto)s,
                 %(descripcion)s, %(lote)s, %(cantidad)s)
            ON CONFLICT DO NOTHING
        """
        execute_batch(cur, sql_entradas, entradas, page_size=100)
        print(f"    ✅ {len(entradas)} entradas importadas")

        # ── Insertar Salidas ─────────────────────────────────────────────────
        print("  Importando inventario_salidas (sin re-calcular stock)...")
        sql_salidas = """
            INSERT INTO inventario_salidas
                (nro_documento, fecha, codigo_producto, descripcion, lote,
                 cantidad, precio_venta, utilidad)
            VALUES
                (%(nro_documento)s, %(fecha)s, %(codigo_producto)s, %(descripcion)s,
                 %(lote)s, %(cantidad)s, %(precio_venta)s, %(utilidad)s)
            ON CONFLICT DO NOTHING
        """
        execute_batch(cur, sql_salidas, salidas, page_size=100)
        print(f"    ✅ {len(salidas)} salidas importadas")

        cur.execute("SET session_replication_role = DEFAULT")  # reactiva triggers
        conn.commit()
        print("\n🎉 Importación completada exitosamente.")

    except Exception as e:
        conn.rollback()
        print(f"\n❌ Error: {e}")
        raise
    finally:
        cur.close()
        conn.close()

# ── Verificación post-importación ────────────────────────────────────────────
def verificar(db_cfg):
    conn = psycopg2.connect(**db_cfg)
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*), SUM(stock_actual), SUM(importe_inventario_proveedor) FROM inventario_stock")
    total, unidades, valor = cur.fetchone()
    cur.execute("SELECT COUNT(*) FROM inventario_entradas")
    ent = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM inventario_salidas")
    sal = cur.fetchone()[0]
    cur.close()
    conn.close()

    print(f"\n📊 VERIFICACIÓN EN BASE DE DATOS:")
    print(f"   inventario_stock   : {total:,} productos | {unidades:,} unidades | ${valor:,.0f} en inventario")
    print(f"   inventario_entradas: {ent:,} registros")
    print(f"   inventario_salidas : {sal:,} registros")

# ── Main ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Importar ControlInventario 2026.xlsx a PostgreSQL")
    parser.add_argument("--excel",   default=DEFAULT_EXCEL, help="Ruta al archivo Excel")
    parser.add_argument("--host",    default=DEFAULT_DB["host"])
    parser.add_argument("--port",    default=DEFAULT_DB["port"], type=int)
    parser.add_argument("--db",      default=DEFAULT_DB["dbname"])
    parser.add_argument("--user",    default=DEFAULT_DB["user"])
    parser.add_argument("--pwd",     default=DEFAULT_DB["password"])
    parser.add_argument("--sslmode",  default=None, help="SSL mode (require, disable, etc.)")
    parser.add_argument("--hostaddr", default=None, help="IP directa para bypasear DNS (ej. Supabase)")
    args = parser.parse_args()

    db_cfg = {
        "host": args.host, "port": args.port,
        "dbname": args.db, "user": args.user, "password": args.pwd,
    }
    if args.sslmode:
        db_cfg["sslmode"] = args.sslmode
    if args.hostaddr:
        db_cfg["hostaddr"] = args.hostaddr

    if not Path(args.excel).exists():
        print(f"❌ Archivo no encontrado: {args.excel}")
        sys.exit(1)

    stock, entradas, salidas = leer_excel(args.excel)
    importar(stock, entradas, salidas, db_cfg)
    verificar(db_cfg)
