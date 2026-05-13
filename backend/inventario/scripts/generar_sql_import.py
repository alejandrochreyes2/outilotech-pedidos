"""Genera inventario_datos_import.sql desde ControlInventario 2026.xlsx"""
import openpyxl
from datetime import date, datetime
from pathlib import Path

def s(v, d=""):
    return str(v).strip() if v is not None else d

def i(v):
    try: return int(v) if v is not None else 0
    except: return 0

def f(v):
    try: return float(v) if v is not None else 0.0
    except: return 0.0

def q(v):
    if isinstance(v, str):
        return v.replace("'", "''")
    return str(v)

def fdate(v):
    if isinstance(v, datetime): return str(v.date())
    if isinstance(v, date): return str(v)
    return str(date.today())

xlsx = r"C:\Users\user\Downloads\ControlInventario 2026.xlsx"
out  = r"C:\Users\user\Downloads\inventario_datos_import.sql"

wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
lines = []
lines.append("-- Importar inventario a produccion (ejecutar en DBeaver)")
lines.append("SET session_replication_role = replica;")
lines.append("TRUNCATE TABLE inventario_salidas, inventario_entradas, inventario_stock RESTART IDENTITY CASCADE;")
lines.append("")

ws = wb["Stock"]
count_stock = 0
for row in ws.iter_rows(min_row=5, values_only=True):
    cod = s(row[0])
    if not cod: continue
    desc = q(s(row[1]) or cod)
    lote = q(s(row[2]) or cod)
    ent   = i(row[3])
    sal   = i(row[4])
    stock = i(row[5])
    costo = f(row[6])
    pventa = f(row[7])
    importe = f(row[8])
    venta = f(row[10]) if len(row) > 10 and row[10] is not None else 0.0
    lines.append(
        f"INSERT INTO inventario_stock "
        f"(codigo_producto,descripcion,lote,entradas,salidas,stock_actual,"
        f"costo_unitario,precio_venta,importe_inventario_proveedor,venta) VALUES "
        f"('{q(cod.upper())}','{desc}','{lote}',{ent},{sal},{stock},"
        f"{costo},{pventa},{importe},{venta}) "
        f"ON CONFLICT (codigo_producto) DO UPDATE SET "
        f"descripcion=EXCLUDED.descripcion,lote=EXCLUDED.lote,"
        f"entradas=EXCLUDED.entradas,salidas=EXCLUDED.salidas,"
        f"stock_actual=EXCLUDED.stock_actual,costo_unitario=EXCLUDED.costo_unitario,"
        f"precio_venta=EXCLUDED.precio_venta,"
        f"importe_inventario_proveedor=EXCLUDED.importe_inventario_proveedor,"
        f"venta=EXCLUDED.venta,updated_at=NOW();"
    )
    count_stock += 1

lines.append("")
ws2 = wb["Entradas"]
count_ent = 0
for row in ws2.iter_rows(min_row=4, values_only=True):
    cod = s(row[2])
    if not cod: continue
    nro  = i(row[0]) or 1
    fecha = fdate(row[1])
    desc  = q(s(row[3]))
    lote  = q(s(row[4]))
    cant  = i(row[5]) or 1
    lines.append(
        f"INSERT INTO inventario_entradas "
        f"(nro_documento,fecha,codigo_producto,descripcion,lote,cantidad) VALUES "
        f"({nro},'{fecha}','{q(cod.upper())}','{desc}','{lote}',{cant}) "
        f"ON CONFLICT DO NOTHING;"
    )
    count_ent += 1

lines.append("")
ws3 = wb["Salidas"]
count_sal = 0
for row in ws3.iter_rows(min_row=4, values_only=True):
    cod = s(row[2])
    if not cod or cod == "-": continue
    nro   = i(row[0]) or 1
    fecha = fdate(row[1])
    desc  = q(s(row[3]))
    lote  = q(s(row[4]))
    cant  = i(row[5]) or 1
    pventa = f(row[6])
    util  = f(row[7])
    lines.append(
        f"INSERT INTO inventario_salidas "
        f"(nro_documento,fecha,codigo_producto,descripcion,lote,cantidad,precio_venta,utilidad) VALUES "
        f"({nro},'{fecha}','{q(cod.upper())}','{desc}','{lote}',{cant},{pventa},{util}) "
        f"ON CONFLICT DO NOTHING;"
    )
    count_sal += 1

lines.append("")
lines.append("SET session_replication_role = DEFAULT;")
lines.append("SELECT COUNT(*) AS productos, SUM(stock_actual) AS unidades FROM inventario_stock;")

wb.close()
Path(out).write_text("\n".join(lines), encoding="utf-8")
print(f"Generado: {out}")
print(f"  Stock   : {count_stock} productos")
print(f"  Entradas: {count_ent} registros")
print(f"  Salidas : {count_sal} registros")
print(f"  Total lineas SQL: {len(lines)}")
